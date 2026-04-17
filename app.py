import subprocess
import sys
import os

req_file = os.path.join(os.path.dirname(__file__), "requirements.txt")
if os.path.exists(req_file):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", req_file, "-q"])

import re
import io
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import streamlit as st

PATTERNS = {
    "Employee Code": [
        r"Employee\s*Number[:\s]+(\d+)",
        r"Employee\s*Code[:\s]+(\d+)",
        r"Emp\.?\s*No\.?[:\s]+(\d+)",
        r"Ref:\s*\(per\)\s*/\s*file\s+(\d+)",
        r"EMP\s*#\s*(\d+)",
    ],
    "Designation": [
        r"Designation[:\s]+(.+)",
    ],
    "Grade": [
        r"Grade[:\s]+(GRD\.[A-Z0-9\.]+)",
        r"Grade[:\s]+([A-Z0-9\-\.]+)",
    ],
    "Old Gross Salary": [
        r"remuneration\s*is\s*being\s*revised\s*from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)\s*to",
        r"(?:gross\s*)?salary\s*(?:is\s*being\s*)?revised\s*from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)\s*to",
        r"revised\s*from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)\s*to",
        r"raising\s*your\s*[Gg]ross\s*[Ss]alary\s*from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)",
        r"[Gg]ross\s*[Ss]alary\s*from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)\s*[\/\-]*\s*\(",
        r"[Gg]ross\s*[Ss]alary\s*from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)\s*to",
        r"from\s*(?:(?:PKR\.?\s*)?(?:Rs\.?|₨)\s*|PKR\.?\s*)([\d,\.]+)\s*[\/\-]*\s*\(",
    ],
    "New Gross Salary": [
        r"(?:remuneration\s*is\s*being\s*revised\s*from\s*[₨Rs\.PKR\s]+[\d,\.]+\s*to\s*PKR\.?\s*)([\d,\.]+)",
        r"(?:revised\s*from\s*[₨Rs\.PKR\s]+[\d,\.]+\s*to\s*PKR\.?\s*)([\d,\.]+)",
        r"(?:remuneration\s*is\s*being\s*revised\s*from\s*[₨Rs\.PKR\s]+[\d,\.]+\s*to\s*Rs\.?\s*)([\d,\.]+)",
        r"(?:revised\s*from\s*[₨Rs\.PKR\s]+[\d,\.]+\s*to\s*Rs\.?\s*)([\d,\.]+)",
        r"(?:to\s*PKR\.?\s*)([\d,\.]+)",
        r"to\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)\s*[\/\-]*\s*\([^)]*01\.\d{2}\.\d{4}",   # "to Rs. 76,175 /- (as on 01.04.2026)"
        r"(?:gross\s*salary\s*from\s*(?:Rs\.?|₨|PKR\.?)\s*[\d,\.]+\s*[\/\-]*\s*to\s*Rs\.?\s*)([\d,\.]+)",
        r"(?:raising\s*your\s*gross\s*salary\s*to\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:gross\s*salary\s*to\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:gross\s*salary\s*in\s*your\s*new\s*grade\s*will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:revised\s*gross\s*salary\s*will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:monthly\s*(?:gross\s*)?salary\s*will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:new\s*gross\s*salary\s*will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:gross\s*salary\s*will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:salary\s*will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:will\s*be\s*Rs\.?)\s*[₨]?\s*([\d,\.]+)",
        r"(?:to\s*Rs\.?\s*[₨]?\s*)([\d,\.]+)\s*[\/\-]*\s*per\s*month",
        r"(?:to\s*Rs\.?\s*)([\d,\.]+)(?:\s*[\/\-]|\s*,|\s+with)",
    ],
    "1st Increment Amount": [
        r"increase\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)\s*[\/\-]*\s*with\s*effect\s*from",
        r"increase\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)\s*[\/\-]*\s*w\.?e\.?f",
        r"(?:CBA|agreement|union)[^.]*?(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)",
    ],
    "2nd Increment Amount": [
        r"increment\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)\s*[\/\-]*\s*[Pp]er\s*month",
        r"grant\s*an\s*increment\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)",
        r"management[^.]*?increment\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*([\d,\.]+)",
    ],
    "Difference": [
        r"(?:gross\s*)?increase\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*(-?[\d,\.]+)\s*[\/\-]*",
        r"increment\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*(-?[\d,\.]+)\s*[\/\-]*",
        r"raise\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*(-?[\d,\.]+)\s*[\/\-]*",
        r"salary\s*increase\s*(?:of|:)\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*(-?[\d,\.]+)\s*[\/\-]*",
        r"revision\s*of\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*(-?[\d,\.]+)\s*[\/\-]*",
        r"increment\s*(?:amount\s*)?(?:is|:)\s*(?:Rs\.?|PKR\.?)\s*[₨]?\s*(-?[\d,\.]+)\s*[\/\-]*",
    ],
    "Date Effective From": [
        r"effective\s*from\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
        r"effective\s*from\s+(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})",
        r"with\s*effect\s*from\s+([A-Za-z]+\s+\d{1,2},?\s*\d{4})",
        r"with\s*effect\s*from\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        r"with\s*effect\s*from\s+(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})",
        r"effective\s*date[:\s]+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
        r"effective\s*date[:\s]+(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})",
        r"w\.?e\.?f\.?\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})",
        r"w\.?e\.?f\.?\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})",
    ],
}

def clean_number(value):
    return value.replace(",", "").rstrip(".") if value else value

def extract_fields(text):
    results = {}
    for field, patterns in PATTERNS.items():
        matched = ""
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                matched = match.group(1).strip()
                break
        results[field] = matched

    # Calculate Difference from New - Old if not found in text
    if not results.get("Difference"):
        try:
            new_sal = float(clean_number(results["New Gross Salary"]))
            old_sal = float(clean_number(results["Old Gross Salary"]))
            results["Difference"] = f"{new_sal - old_sal:,.2f}"
        except Exception:
            pass

    return results

def process_pdf(file_bytes, filename):
    results = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            fields = extract_fields(text)
            label = filename if total == 1 else f"{filename} — Page {i}"
            results.append((label, fields))
    return results

def build_excel(data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Data"

    headers = [
        "Source", "Employee Code", "Designation", "Grade",
        "Old Gross Salary", "1st Increment Amount", "2nd Increment Amount",
        "New Gross Salary", "Difference", "Date Effective From"
    ]

    header_fill = PatternFill("solid", start_color="2F5496", end_color="2F5496")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_font = Font(name="Arial", size=10)
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ── STREAMLIT UI ────────────────────────────────────────────────
st.set_page_config(page_title="Salary Letter Extractor", page_icon="📄", layout="centered")

st.title("📄 Salary Letter Extractor")
st.markdown("Upload one or more PDF files and click **Extract** to generate an Excel file.")

uploaded_files = st.file_uploader(
    "Upload PDF file(s)",
    type=["pdf"],
    accept_multiple_files=True
)

if uploaded_files:
    if st.button("Extract & Generate Excel", type="primary"):
        all_rows = []
        total_pages = 0
        not_found_count = 0

        with st.spinner("Processing..."):
            for uploaded_file in uploaded_files:
                file_bytes = uploaded_file.read()
                page_results = process_pdf(file_bytes, uploaded_file.name)
                total_pages += len(page_results)

                for label, fields in page_results:
                    row = [
                        label,
                        fields.get("Employee Code", ""),
                        fields.get("Designation", ""),
                        fields.get("Grade", ""),
                        clean_number(fields.get("Old Gross Salary", "")),
                        clean_number(fields.get("1st Increment Amount", "")),
                        clean_number(fields.get("2nd Increment Amount", "")),
                        clean_number(fields.get("New Gross Salary", "")),
                        clean_number(fields.get("Difference", "")),
                        fields.get("Date Effective From", ""),
                    ]
                    all_rows.append(row)
                    not_found_count += sum(1 for v in fields.values() if not v)

        if not all_rows:
            st.error("No data could be extracted from the uploaded files.")
        else:
            fully_extracted = sum(1 for row in all_rows if "" not in row)

            col1, col2, col3 = st.columns(3)
            col1.metric("Letters Processed", total_pages)
            col2.metric("Fully Extracted", fully_extracted)
            col3.metric("Fields Not Found", not_found_count)

            excel_buffer = build_excel(all_rows)
            base_name = os.path.splitext(uploaded_files[0].name)[0]
            output_name = f"SalaryData-{base_name}.xlsx"

            st.download_button(
                label="⬇️ Download Excel",
                data=excel_buffer,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
